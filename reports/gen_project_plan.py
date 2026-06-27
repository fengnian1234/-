"""生成项目开发计划 Excel 文档"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════
# 样式
# ═══════════════════════════════════════════════
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1A232D", end_color="1A232D", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1A232D")
SECTION_FONT = Font(name="微软雅黑", size=12, bold=True, color="2C3E50")
NORMAL_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", size=10, bold=True)
SMALL_FONT = Font(name="微软雅黑", size=9, color="666666")

DONE_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
IN_PROGRESS_FILL = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
PENDING_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
PHASE1_FILL = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
PHASE2_FILL = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
PHASE3_FILL = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='D5D8DC'),
    right=Side(style='thin', color='D5D8DC'),
    top=Side(style='thin', color='D5D8DC'),
    bottom=Side(style='thin', color='D5D8DC'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)


def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = thin_border


def style_row(ws, row, ncols, fill=None):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = NORMAL_FONT
        cell.alignment = LEFT if col > 1 else CENTER
        cell.border = thin_border
        if fill:
            cell.fill = fill


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════════
# Sheet 1: 项目总览
# ═══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "项目总览"

overview = [
    ("云上归墅 — 微信公众号 AI 客服系统", TITLE_FONT),
    ("", None),
    ("项目阶段", "未落成（项目规划提案阶段，面向第一个客户）"),
    ("开发周期", "2025年6月 — 至今"),
    ("当前版本", "v3.3"),
    ("", None),
    ("【代码规模】", SECTION_FONT),
    ("Python 源代码", "7,933 行 / 25 个文件"),
    ("HTML 模板 + 预览", "42 个（Jinja2 模板 18 个 + 纯HTML预览 24 个）"),
    ("CSS 样式表", "2,860 行（static/css/style.css）"),
    ("数据库表", "15 张（SQLAlchemy ORM）"),
    ("Flask 路由", "28 条"),
    ("Docker 服务", "3 个（Flask + PostgreSQL 15 + Nginx）"),
    ("", None),
    ("【测试 & 质量】", SECTION_FONT),
    ("单元测试", "30 个 / 9 个测试类 / 28 通过 2 预存失败"),
    ("Git 提交", "26 次（v2.5 → v3.3，含完整版本历史）"),
    ("日志系统", "RotatingFileHandler + error.log 独立输出"),
    ("安全审计", "已完成：PII 脱敏 / XSS 防护 / SECRET_KEY 自动生成"),
    ("", None),
    ("【技术栈】", SECTION_FONT),
    ("后端", "Python 3.12 + Flask 3.1 + SQLAlchemy 2.0"),
    ("AI", "DeepSeek API（Anthropic 兼容接口，model=deepseek-chat）"),
    ("微信", "wechatpy 2.1（消息收发 / 事件处理 / 菜单 / 模板消息）"),
    ("数据库", "开发 SQLite / 生产 PostgreSQL 15"),
    ("部署", "Docker Compose（Gunicorn 4 workers + Nginx 反向代理）"),
    ("监控搜索", "opencli（携程/点评/小红书/微博/知乎原生适配器）+ WebSearch Bing 兜底"),
]

for i, (k, v) in enumerate(overview, 1):
    if k == "":
        continue
    ws1.cell(row=i, column=1, value=k).font = v if isinstance(v, Font) else BOLD_FONT
    ws1.cell(row=i, column=2, value=v if not isinstance(v, Font) else k)
    if isinstance(v, Font):
        ws1.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)

ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 65

# ═══════════════════════════════════════════════
# Sheet 2: 已完成功能清单
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet("已完成功能")

headers2 = ["模块", "功能", "版本", "备注"]
for j, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=j, value=h)
style_header(ws2, 1, len(headers2))

completed = [
    # 微信消息
    ("微信消息", "消息收发 + XML 解析 + 签名验证", "v2.1", "wechat.py 完整实现"),
    ("微信消息", "30+ 关键词正则路由（房型/菜单/攻略/服务/预订等）", "v2.5", "支持正则捕获组"),
    ("微信消息", "微信事件处理（关注/取消/扫码/菜单点击）", "v2.8", ""),
    # AI
    ("AI 对话", "四层权限体系（travel_advisor / pre_arrival / guest_butler / post_stay）", "v3.2", "4 套独立 System Prompt，含具体规则"),
    ("AI 对话", "本地数据注入（房型/菜单/路线/美食/文档自动注入 Prompt）", "v3.1", "local_data 优先级最高"),
    ("AI 对话", "请求间隔控制（每 3 秒最多 1 次 AI 请求）", "v3.2.4", "线程锁 + 时间戳"),
    ("AI 对话", "输入校验（长度截断 500 字 / 去重 / 控制符清理）", "v3.2.2", "_validate_and_sanitize()"),
    ("AI 对话", "截断继续生成（stop_reason=max_tokens 检测 + 回复「继续」续写）", "v3.2.5", "DeepSeek 原生 stop_reason 字段"),
    ("AI 对话", "DeepSeek API 对接（Anthropic 兼容接口）", "v2.0", "base_url + model 可配置"),
    ("AI 对话", "对话持久化（MessageLog 表 / 最近10轮上下文加载）", "v3.2", "_save_conversation / _load_conversation"),
    # 预订
    ("预订管理", "预订确认 + 状态流转（confirmed → checked_in → checked_out → completed）", "v2.0", ""),
    ("预订管理", "房间码共享机制（6位随机码 / 同住人绑定 / 主动告知）", "v3.2", "RoomGuest 表 / 去重"),
    ("预订管理", "续住延期（解析 check_out_date + 增加天数）", "v3.2", "AI 权限保持不降级"),
    ("预订管理", "离店后好评推送（退房 30 分钟后 / 6 平台评价链接）", "v3.2", "REVIEW_PLATFORMS 配置"),
    # 积分
    ("积分体系", "签到 / 消费 / 评价积分 + 银/金/钻石三级会员", "v2.13", "含会员等级进度条"),
    ("积分体系", "积分兑换商品 + 生日 1.5 倍奖励", "v2.13", ""),
    # 业务
    ("房型展示", "8 种房型 + 携程真实数据 + 66 张实拍图 + 详情轮播", "v2.5", "图片从携程 CDN 下载"),
    ("菜单系统", "4 分类 21 品咖啡简餐 + 饮品定制弹窗 + 推荐 + 下单", "v2.14", "三山二水咖啡"),
    ("旅游攻略", "5 条游玩路线 + 8 个美食推荐 + 位置/交通地图链接", "v2.5", "含高德/腾讯地图"),
    ("快捷服务", "21 项服务 + 自动创建通知给员工", "v2.13", "模糊匹配服务名"),
    ("员工看板", "实时仪表盘 + 5s 自动刷新 + 声音提醒 + 待处理列表", "v2.8", "暗色模式"),
    # 监控
    ("口碑监控", "6 平台搜索收集 + 情感分析 + URL 去重存储", "v3.3", "opencli + WebSearch 双引擎"),
    ("口碑监控", "评分提取 + 评价数统计 + 汇总报表生成", "v3.3", ""),
    ("口碑监控", "DOCX 周报生成（封面 + 数据表 + 趋势分析 + 平台截图区）", "v2.0", "python-docx 中式排版"),
    # 前端
    ("前端页面", "三季 UI 主题系统（春夏·秋·冬 / CSS 变量 / data-theme）", "v2.14", "全站覆盖"),
    ("前端页面", "深色/浅色模式切换", "v2.8", "data-theme=dark 选择器"),
    ("前端页面", "响应式布局（手机 / 平板 / 桌面）", "v2.5", ""),
    ("前端页面", "主页（Hero + 快捷入口 + 房型预览 + 评价轮播 + FAQ 手风琴）", "v2.11", "动态季节文案"),
    ("前端页面", "微信模拟器（4 种 AI 模式 + 快捷词 + 批量测试 + 事件模拟）", "v3.2", "调试工具 / 桌面+移动端"),
    # 基础设施
    ("基础设施", "Docker 三服务部署（Flask + PostgreSQL + Nginx + 健康检查）", "v3.2", "生产可用"),
    ("基础设施", "统一日志系统（RotatingFileHandler 10MB×30 + error.log 独立）", "v3.2", "7 个便捷函数"),
    ("基础设施", "单元测试 30 个（9 类：房型/菜单/攻略/服务/预订/积分/订单/通知/AI/日志）", "v3.2", "28/30 通过"),
    ("基础设施", "API 文档页 /docs + 数据库 Schema 文档", "v3.2", "28 端点 + 15 表"),
    ("基础设施", "安全加固（PII 脱敏 / XSS 防护 / SECRET_KEY 自动生成 / DEBUG 默认关）", "v3.2.1", ""),
    ("基础设施", "代码质量收尾（print → logger / 异常加日志 / og-image / 依赖补齐）", "v3.2.6", ""),
    ("基础设施", "AnySearch 移除 → opencli + WebSearch 双引擎替代", "v3.3", "无限额免费搜索"),
]

for i, row in enumerate(completed, 2):
    for j, val in enumerate(row, 1):
        ws2.cell(row=i, column=j, value=val)
    style_row(ws2, i, len(headers2), DONE_FILL)

set_col_widths(ws2, [14, 58, 10, 42])

# ═══════════════════════════════════════════════
# Sheet 3: 后续开发计划
# ═══════════════════════════════════════════════
ws3 = wb.create_sheet("后续开发计划")

headers3 = ["优先级", "阶段", "模块", "任务", "说明", "依赖条件", "预计工期"]
for j, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=j, value=h)
style_header(ws3, 1, len(headers3))

roadmap = [
    # P0 — 上线前必须完成
    ("P0", "上线前", "微信支付", "真实微信支付 API 集成",
     "替换 menu.py 中的 wx_prepay_* 占位符，接入统一下单 API + 支付回调处理",
     "微信商户号 + API 密钥", "2 天"),
    ("P0", "上线前", "微信客服消息", "客服消息 API 集成",
     "解除 wechat.py:551 注释，实现好评推送 / 服务通知 / 退房提醒主动发送",
     "微信公众号 AppID + AppSecret", "1 天"),
    ("P0", "上线前", "微信公众号", "公众号注册 + 认证 + 服务器配置",
     "注册服务号(企业主体) → 微信认证(300元/年) → 配置 URL / Token / IP 白名单 / 菜单",
     "营业执照", "2-4 周"),
    ("P0", "上线前", "云服务器", "服务器采购 + Docker 部署",
     "阿里云/腾讯云 2c4g(约100元/月) → 安装 Docker CE → git clone → docker compose up -d",
     "域名 + 服务器", "1 天"),
    ("P0", "上线前", "域名 + SSL", "域名注册 + 备案 + HTTPS 证书",
     "注册域名(yunshang.cc等) → 工信部备案(约20工作日) → Let's Encrypt SSL → Nginx HTTPS",
     "营业执照", "3-4 周"),
    ("P0", "上线前", "opencli", "各平台官方账号注册 + Cookie 持久化",
     "注册大众点评商家号 / 小红书官方号 / 微博 → Chrome 独立配置 → opencli 登录持久化",
     "营业执照 + 手机号", "1 天"),
    ("P0", "上线前", "真实数据", "替换种子数据为真实运营数据",
     "导入民宿真实房型 / 菜单 / 价格 / 图片 / 路线 / 美食 / 员工信息",
     "民宿运营方提供", "1-2 天"),
    ("P0", "上线前", "全链路测试", "端到端测试全部功能",
     "关注 → 菜单 → 关键词回复 → AI 对话 → 预订 → 支付 → 退房 → 好评推送 → 周报",
     "以上全部完成", "1 天"),
    # P1 — 上线后优先
    ("P1", "上线后", "微信支付", "微信内 H5 支付完整流程",
     "JSAPI 支付 → 统一下单 → 唤起收银台 → 支付结果通知",
     "P0 微信支付基础完成", "2 天"),
    ("P1", "上线后", "微信模板消息", "模板消息申请 + 发送",
     "提交预订确认 / 服务通知 / 好评提醒模板审核 → 对接发送接口",
     "公众号认证完成", "1-2 周"),
    ("P1", "上线后", "数据备份", "PostgreSQL 自动备份",
     "pg_dump 定时任务 → 备份到 OSS/本地 → 保留最近 7 天",
     "服务器", "0.5 天"),
    ("P1", "上线后", "监控告警", "生产监控 + 日志告警",
     "Uptime 监控 (/health) → ERROR 关键词 → 钉钉/企业微信通知",
     "服务器 + 企微群", "1 天"),
    ("P1", "上线后", "Cron 周报", "修正 Cron 周报为 opencli 搜索",
     "Claude Code Cron: 每周一 opencli 收集各平台数据 → 存入 DB → 生成 DOCX",
     "opencli Cookie 配置完成", "0.5 天"),
    ("P1", "上线后", "CI/CD", "GitHub Actions 自动测试 + 部署",
     "push → pytest → docker build → 自动部署到服务器",
     "服务器 SSH 配置", "1 天"),
    ("P1", "上线后", "员工培训", "民宿员工操作培训",
     "培训使用员工看板 / 订单管理 / 服务通知处理 / 数据查询",
     "系统部署完成", "1 天"),
    # P2 — 持续优化
    ("P2", "持续优化", "AI", "AI 对话质量持续优化",
     "根据真实用户反馈调整 System Prompt / 添加 FAQ 快速匹配 / 方言支持",
     "运营数据积累", "持续"),
    ("P2", "持续优化", "PMS 对接", "民宿管理系统对接",
     "与民宿 PMS 系统对接实时房态 / 预订同步 / 自动入住/退房",
     "PMS 系统 API", "3-5 天"),
    ("P2", "持续优化", "数据库", "数据库迁移方案 Alembic",
     "生产环境 Schema 变更用 Alembic 版本管理，替代 create_all()",
     "", "1 天"),
    ("P2", "持续优化", "多语言", "英文/日文客服支持",
     "针对外国游客的 AI 翻译 + 多语言回复",
     "用户需求确认", "2-3 天"),
    ("P2", "持续优化", "多店", "多门店架构预留",
     "当前单店设计，预留多店切换 / 数据隔离接口",
     "业务扩展需求", "待评估"),
    ("P2", "持续优化", "数据分析", "运营数据分析仪表盘",
     "入住率 / 客单价 / 好评率 / 平台流量趋势 / AI 对话统计",
     "数据积累 3 个月+", "3-5 天"),
    ("P2", "持续优化", "小红书", "小红书自动发布笔记",
     "opencli xiaohongshu publish → 定时发布民宿推广内容",
     "小红书官方号 + Cookie", "1 天"),
    ("P2", "持续优化", "抖音", "抖音内容监控 + 发布",
     "通过 WebSearch 监控抖音提及 + 评估 opencli douyin publish",
     "抖音企业号", "2 天"),
]

for i, row in enumerate(roadmap, 2):
    for j, val in enumerate(row, 1):
        ws3.cell(row=i, column=j, value=val)
    priority = row[0]
    fill = PENDING_FILL if priority == "P0" else (PHASE1_FILL if priority == "P1" else PHASE3_FILL)
    style_row(ws3, i, len(headers3), fill)

set_col_widths(ws3, [8, 12, 16, 34, 52, 26, 12])

# ═══════════════════════════════════════════════
# Sheet 4: 落地部署流程
# ═══════════════════════════════════════════════
ws4 = wb.create_sheet("落地部署流程")

headers4 = ["步骤", "阶段", "事项", "具体操作", "产出物", "预计耗时"]
for j, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=j, value=h)
style_header(ws4, 1, len(headers4))

deploy = [
    # 第 0 阶段：现在可做的技术准备
    ("0.1", "第0阶段：技术准备", "opencli Cookie 配置",
     "注册各平台官方号 → Chrome 独立配置目录 → opencli browser 登录 → --site-session persistent 持久化",
     "所有平台 Cookie 就绪", "1 天"),
    ("0.2", "第0阶段：技术准备", "真实数据准备清单",
     "收集民宿真实：房型名称/面积/床型/价格 + 菜单品类/价格 + 图片 + 周边路线/美食",
     "数据清单 Excel", "1-2 天"),
    ("0.3", "第0阶段：技术准备", "微信支付接口开发",
     "services/menu.py: 替换 _get_prepay_id() 占位 → wechatpy 统一下单 API → wechat/pay-notify 回调处理",
     "支付功能代码就绪", "2 天"),
    ("0.4", "第0阶段：技术准备", "客服消息接口开发",
     "wechat.py: 解除 send_wechat_customer_message 注释 → 对接客服消息 API → 文本/图文消息发送",
     "主动推送代码就绪", "1 天"),
    # 第 1 阶段：账号 & 基础设施（可与第0阶段并行）
    ("1.1", "第1阶段：账号 & 基础设施", "微信公众号注册",
     "注册服务号(企业主体) → 提交营业执照 → 微信认证(300元/年) → 获取 AppID + AppSecret",
     "AppID / AppSecret / Token", "1-2 周"),
    ("1.2", "第1阶段：账号 & 基础设施", "微信支付商户号申请",
     "提交营业执照 → 商户号审核 → 获取 MCH_ID + API 密钥 MCH_KEY",
     "商户号 / API 密钥", "1-2 周"),
    ("1.3", "第1阶段：账号 & 基础设施", "域名注册 + 备案",
     "注册域名(如 yunshang.cc) → 提交备案资料 → 等待管局审核(约20工作日) → 备案号",
     "域名 + 备案号", "3-4 周"),
    ("1.4", "第1阶段：账号 & 基础设施", "云服务器采购",
     "阿里云/腾讯云 2c4g(约100元/月) → 安装 Docker CE + Docker Compose → 配置安全组(80/443/5000)",
     "服务器 IP + SSH 登录", "1 天"),
    # 第 2 阶段：部署上线
    ("2.1", "第2阶段：部署上线", "生产环境变量配置",
     "复制 .env.prod → .env → 填入 WECHAT_*/ANTHROPIC_API_KEY/DATABASE_URL/SECRET_KEY(openssl rand -hex 32)",
     ".env 生产配置", "0.5 天"),
    ("2.2", "第2阶段：部署上线", "Docker 部署",
     "git clone → docker compose up -d → 验证 PostgreSQL/Flask(Gunicorn 4 workers)/Nginx 三服务",
     "服务正常运行", "1 小时"),
    ("2.3", "第2阶段：部署上线", "SSL 证书 + HTTPS",
     "certbot 申请 Let's Encrypt 证书 → Nginx HTTPS 配置 → certbot renew cron 自动续期",
     "HTTPS 可用", "1 小时"),
    ("2.4", "第2阶段：部署上线", "微信服务器接入验证",
     "公众号后台配置 URL: https://域名/wechat → Token 验证 → GET 请求签名校验通过",
     "微信消息收发正常", "1 小时"),
    ("2.5", "第2阶段：部署上线", "端到端全链路测试",
     "关注公众号 → 菜单点击 → 关键词回复 → AI 对话 → 预订流程 → 支付 → 入住 → 离店 → 好评推送",
     "全链路测试报告", "1 天"),
    # 第 3 阶段：运营完善
    ("3.1", "第3阶段：运营完善", "微信菜单配置",
     "通过 API 或后台配置公众号底部菜单：房型/菜单/攻略/服务/客服",
     "菜单可用", "0.5 天"),
    ("3.2", "第3阶段：运营完善", "模板消息申请",
     "提交预订确认/服务通知/好评提醒模板 → 微信审核 → 获取模板 ID",
     "模板 ID", "1-2 周"),
    ("3.3", "第3阶段：运营完善", "Cron 周报配置",
     "Claude Code Cron: 每周一 9:57 → opencli 收集各平台数据 → 存入 DB → 生成 DOCX → 保存到 reports/",
     "自动周报运行", "0.5 天"),
    ("3.4", "第3阶段：运营完善", "监控告警配置",
     "UptimeRobot 监控 /health → 日志 ERROR 关键词 → 钉钉/企微 Webhook 通知",
     "告警可用", "0.5 天"),
    ("3.5", "第3阶段：运营完善", "日常运营培训",
     "培训民宿员工：员工看板使用 / 订单管理 / 服务请求处理 / 数据查询 / 常见问题处理",
     "员工操作手册", "1 天"),
]

for i, row in enumerate(deploy, 2):
    for j, val in enumerate(row, 1):
        ws4.cell(row=i, column=j, value=val)
    phase = row[1]
    if "第0" in phase:
        fill = PENDING_FILL
    elif "第1" in phase:
        fill = PHASE1_FILL
    elif "第2" in phase:
        fill = PHASE2_FILL
    else:
        fill = PHASE3_FILL
    style_row(ws4, i, len(headers4), fill)

set_col_widths(ws4, [8, 22, 24, 62, 24, 12])

# ═══════════════════════════════════════════════
# Sheet 5: 技术架构
# ═══════════════════════════════════════════════
ws5 = wb.create_sheet("技术架构")

headers5 = ["层次", "文件", "行数", "职责", "关键特性"]
for j, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=j, value=h)
style_header(ws5, 1, len(headers5))

arch = [
    ("路由/控制器", "app.py", "880", "28 条 Flask 路由 + 应用初始化", "限流 / 错误处理 / 模拟器 API / 健康检查"),
    ("路由/控制器", "wechat.py", "563", "微信消息处理 + 30+ 关键词正则路由", "四层 AI 降级 / 事件处理 / 好评推送调度"),
    ("业务服务", "services/ai.py", "658", "AI 对话（四层权限 + 本地数据注入 + 并发控制）", "截断继续 / 对话持久化 / 间隔控制 / 输入校验"),
    ("业务服务", "services/booking.py", "342", "预订管理（CRUD + 状态流转 + 房间码）", "RoomGuest 共享 / 续住 / 离店 / 好评推送查询"),
    ("业务服务", "services/monitor.py", "530", "口碑监控（opencli + WebSearch 双引擎）", "6 平台搜索 / 情感分析 / URL 去重 / 评分解析"),
    ("业务服务", "services/report.py", "348", "DOCX 周报生成", "封面 + 数据表 + 趋势分析 + 平台截图区"),
    ("业务服务", "services/menu.py", "254", "菜单查询 + 下单 + 微信支付参数", "4 分类 21 品 / 饮品定制 / 预支付"),
    ("业务服务", "services/travel.py", "244", "游玩路线 + 美食推荐 + 位置", "5 路线 / 8 美食 / 高德腾讯地图链接"),
    ("业务服务", "services/orders.py", "188", "多平台订单聚合", "仪表盘统计 / 14 天房态日历"),
    ("业务服务", "services/points.py", "166", "积分体系 + 会员等级", "签到/消费/评价/兑换/生日加成"),
    ("业务服务", "services/notify.py", "131", "员工通知 + 服务请求队列", "待处理/已确认/已完成/今日统计"),
    ("业务服务", "services/rooms.py", "124", "房型查询 + 格式化输出", "8 房型 / 图文 / 微信格式"),
    ("业务服务", "services/quick.py", "114", "21 项快捷服务", "模糊匹配名称 / 自动创建服务请求"),
    ("业务服务", "services/logger.py", "106", "统一日志系统", "RotatingFileHandler 10MBx30 / error.log / 7 便捷函数"),
    ("数据层", "models.py", "544", "15 张表 SQLAlchemy ORM", "Booking / RoomGuest / MessageLog / PlatformMention"),
    ("数据层", "config.py", "143", "全部配置集中管理", ".env + .env.local 覆盖 / 平台链接 / 民宿信息"),
    ("数据层", "seed_data.py", "640", "种子数据（携程真实数据）", "8 房型 / 21 菜品 / 21 服务 / 5 路线 / 8 美食"),
    ("前端", "templates/", "~3500", "18 个 Jinja2 模板（含 base.html）", "三季主题 / 暗色模式 / 响应式 / 轮播 / 手风琴"),
    ("前端", "preview/", "~2500", "24 个纯 HTML 预览页", "自包含 / 可直接双击打开 / 与模板功能对等"),
    ("前端", "static/css/", "2860", "全站响应式样式表", "CSS 自定义属性 / 三季主题 / 动画 / 模态窗"),
    ("部署", "Dockerfile + compose", "~120", "Python 3.10 + Gunicorn + PostgreSQL + Nginx", "健康检查 / 卷 / 网络 / 入口脚本"),
    ("测试", "tests/", "282", "30 个单元测试 / 9 个测试类", "覆盖全部 services/ 模块"),
]

for i, row in enumerate(arch, 2):
    for j, val in enumerate(row, 1):
        ws5.cell(row=i, column=j, value=val)
    layer = row[0]
    fills = {
        "路由/控制器": PHASE1_FILL,
        "业务服务": DONE_FILL,
        "数据层": PHASE3_FILL,
        "前端": PHASE2_FILL,
        "部署": PENDING_FILL,
        "测试": DONE_FILL,
    }
    style_row(ws5, i, len(headers5), fills.get(layer))

set_col_widths(ws5, [14, 24, 8, 44, 50])

# ═══════════════════════════════════════════════
# Sheet 6: 平台适配器状态
# ═══════════════════════════════════════════════
ws6 = wb.create_sheet("平台适配器状态")

headers6 = ["平台", "opencli 适配器", "搜索命令", "需 Browser", "Cookie 状态", "降级方案", "备注"]
for j, h in enumerate(headers6, 1):
    ws6.cell(row=1, column=j, value=h)
style_header(ws6, 1, len(headers6))

platforms = [
    ("携程", "✅ ctrip", "opencli ctrip search", "否（公开接口）", "无需配置", "—", "民宿已上线携程，搜索为公开能力"),
    ("大众点评", "✅ dianping", "opencli dianping search --city 九江", "是", "待注册商家号", "WebSearch (Bing)", "免费注册商家号即可"),
    ("小红书", "✅ xiaohongshu", "opencli xiaohongshu search", "是", "待注册官方号", "WebSearch (Bing)", "支持 search/note/download/feed"),
    ("微博", "✅ weibo", "opencli weibo search", "是", "待注册官方号", "WebSearch (Bing)", "支持 search/hot/user/comments"),
    ("知乎", "✅ zhihu", "opencli zhihu search", "是", "待注册账号", "WebSearch (Bing)", "支持 search/question/answer"),
    ("美团", "❌ 无适配器", "—", "—", "—", "WebSearch (Bing)", "暂无 opencli 适配器"),
    ("飞猪", "❌ 无适配器", "—", "—", "—", "WebSearch (Bing)", "暂无 opencli 适配器"),
    ("抖音", "⚠️ 仅 publish", "—", "—", "—", "WebSearch (Bing)", "有 publish 适配器，无 search"),
]

for i, row in enumerate(platforms, 2):
    for j, val in enumerate(row, 1):
        ws6.cell(row=i, column=j, value=val)
    has_adapter = row[1].startswith("✅")
    needs_browser = row[3] == "是"
    if has_adapter and not needs_browser:
        fill = DONE_FILL
    elif has_adapter and needs_browser:
        fill = PHASE1_FILL
    else:
        fill = PENDING_FILL
    style_row(ws6, i, len(headers6), fill)

set_col_widths(ws6, [12, 18, 42, 16, 18, 24, 40])

# ═══════════════════════════════════════════════
# 保存
# ═══════════════════════════════════════════════
output = os.path.join(os.path.dirname(os.path.abspath(__file__)), "云上归墅_项目开发计划_20260614.xlsx")
wb.save(output)
print(f"✅ Excel 已保存: {output}")
print(f"共 {len(wb.sheetnames)} 个工作表: {', '.join(wb.sheetnames)}")
