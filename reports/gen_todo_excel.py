#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""生成待办事项清单 Excel"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

wb = openpyxl.Workbook()

# ── Styles ──
header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='2a3d33', end_color='2a3d33', fill_type='solid')
normal_font = Font(name='Microsoft YaHei', size=9)
bold_font = Font(name='Microsoft YaHei', size=9, bold=True)
thin_border = Border(
    left=Side(style='thin', color='cccccc'),
    right=Side(style='thin', color='cccccc'),
    top=Side(style='thin', color='cccccc'),
    bottom=Side(style='thin', color='cccccc'),
)
wrap = Alignment(wrap_text=True, vertical='center')
center = Alignment(horizontal='center', vertical='center')

# Priority colors
p0_fill = PatternFill(start_color='ffe0e0', end_color='ffe0e0', fill_type='solid')
p1_fill = PatternFill(start_color='fff3e0', end_color='fff3e0', fill_type='solid')
p2_fill = PatternFill(start_color='e8f5e9', end_color='e8f5e9', fill_type='solid')
p3_fill = PatternFill(start_color='e3f2fd', end_color='e3f2fd', fill_type='solid')

# ── Data ──
tasks = [
    # (priority, category, task, detail, effort, dependency, file_path)
    ("P0", "🏠 山纪/东林外·基础信息", "山纪真实地址", "config.py BNB_CONFIGS['shanji']['address'] 当前为 'XX路XX号'，需填入真实门牌号", "小", "向山纪店主确认", "config.py:78"),
    ("P0", "🏠 山纪/东林外·基础信息", "东林外真实地址", "config.py BNB_CONFIGS['donglinwai']['address'] 当前为 'XX路XX号'，需填入真实门牌号", "小", "向东林外店主确认", "config.py:91"),
    ("P0", "🏠 山纪/东林外·基础信息", "山纪联系电话", "config.py BNB_CONFIGS['shanji']['phone'] 当前为 '待填写'", "小", "向山纪店主确认", "config.py:79"),
    ("P0", "🏠 山纪/东林外·基础信息", "东林外联系电话", "config.py BNB_CONFIGS['donglinwai']['phone'] 当前为 '待填写'", "小", "向东林外店主确认", "config.py:92"),
    ("P0", "🏠 山纪/东林外·基础信息", "山纪精确坐标", "config.py 中 latitude=29.5600, longitude=115.9850 为估算值，需确认或实地取点", "小", "店主确认/地图取点", "config.py:80-81"),
    ("P0", "🏠 山纪/东林外·基础信息", "东林外精确坐标", "config.py 中 latitude=29.5500, longitude=115.9700 为估算值", "小", "店主确认/地图取点", "config.py:93-94"),
    ("P0", "🏠 山纪/东林外·基础信息", "山纪民宿描述优化", "当前 description='茶园环绕，品茗观山...' 过于简略，需丰富为类似归墅的详细描述", "小", "山纪店主提供特色信息", "config.py:83"),
    ("P0", "🏠 山纪/东林外·基础信息", "东林外民宿描述优化", "当前 description='东林寺旁，禅意疗愈之所...' 过于简略", "小", "东林外店主提供特色信息", "config.py:96"),
    ("P0", "🏠 山纪/东林外·基础信息", "小程序常量同步", "miniapp/utils/constants.js 中 shanji/donglinwai 地址电话同样为占位符，需同步更新", "小", "上述信息确认后", "miniapp/utils/constants.js:25-38"),
    ("P0", "🏠 山纪/东林外·基础信息", "迁移脚本同步", "scripts/migrate_multi_bnb.py 中 shanji/donglinwai 地址电话同样为占位符", "小", "上述信息确认后", "scripts/migrate_multi_bnb.py:64,71"),

    ("P0", "🛏️ 山纪·客房/菜单", "山纪客房数据", "seed_data.py 中完全没有 shanji 的客房 seed 数据。需要：房型名称、价格、面积、床型、设施、景观", "大", "山纪店主提供房型资料", "seed_data.py"),
    ("P0", "🛏️ 山纪·客房/菜单", "山纪菜单数据", "seed_data.py 中没有 shanji 的 menu_items。需要：菜品名称、分类、价格、描述", "中", "山纪店主提供菜单", "seed_data.py"),
    ("P0", "🛏️ 山纪·客房/菜单", "山纪快捷服务", "seed_data.py 中没有 shanji 的 service_requests 分类数据", "小", "山纪支持的服务项目", "seed_data.py"),
    ("P0", "🏕️ 东林外·客房/菜单", "东林外客房数据", "seed_data.py 中完全没有 donglinwai 的客房 seed 数据。需要：房型、价格、面积、床型、设施、景观", "大", "东林外店主提供房型资料", "seed_data.py"),
    ("P0", "🏕️ 东林外·客房/菜单", "东林外菜单数据", "seed_data.py 中没有 donglinwai 的 menu_items。需要：菜品名称、分类、价格、描述", "中", "东林外店主提供菜单", "seed_data.py"),
    ("P0", "🏕️ 东林外·客房/菜单", "东林外快捷服务", "seed_data.py 中没有 donglinwai 的 service_requests 分类数据", "小", "东林外支持的服务项目", "seed_data.py"),

    ("P0", "☁️ 域名与服务器", "购买域名", "确定域名并注册（候选：yunshangguishu.com / yunshangbnb.com）", "小", "无", "外部"),
    ("P0", "☁️ 域名与服务器", "代码中硬编码域名统一", "app.py CORS trusted + 启动信息 + nginx.conf server_name + docker-compose BASE_URL 共 6 处改为读环境变量", "小", "域名确定后", "app.py:46,1225 / nginx.conf:17 / docker-compose.yml:41"),
    ("P0", "☁️ 域名与服务器", "小程序 API 地址环境感知", "miniapp/utils/constants.js API_BASE_URL 硬编码 localhost，需根据 envVersion 自动切换", "小", "域名确定后", "miniapp/utils/constants.js:6"),
    ("P0", "☁️ 域名与服务器", "租用云服务器", "选购配置（2C4G Linux，国内需备案），Docker Compose 一键部署栈已就绪", "中", "域名已注册", "外部"),
    ("P0", "☁️ 域名与服务器", "DNS 解析 A 记录", "将域名解析到服务器 IP", "小", "服务器 IP 已知", "外部"),
    ("P0", "☁️ 域名与服务器", "SSL 证书签发", "Let's Encrypt / acme.sh 自动签发 + 续期，nginx HTTPS 块取消注释", "小", "DNS 解析生效", "nginx.conf:79-90"),
    ("P0", "☁️ 域名与服务器", "ICP 备案", "国内服务器需要 ICP 备案，周期约 2-3 周。现在可以开始准备材料", "大(等待期)", "域名已实名", "外部"),
    ("P0", "☁️ 域名与服务器", "CORS 信任来源动态化", "app.py 中 CORS trusted 列表改为从 BASE_URL 自动派生，而非硬编码", "小", "域名确定后", "app.py:43-47"),

    ("P0", "🔗 微信接入", "微信公众号认证", "年审 300 元，需要营业执照。认证后才能使用客服消息 API", "中(等待期)", "营业执照", "外部"),
    ("P0", "🔗 微信接入", "公众号服务器 URL 配置", "在公众号后台填写 URL: https://域名/wechat，完成 Token 验证", "小", "服务器就绪 + 域名 HTTPS", "公众号后台"),
    ("P0", "🔗 微信接入", "微信客服消息 API 实现", "wechat.py:613 中 send_wechat_customer_message() 调用被注释，需实现 get_access_token() → 调用微信 API 发送消息", "中", "公众号认证通过", "wechat.py:613"),
    ("P0", "🔗 微信接入", "微信支付统一下单 API", "services/menu.py:251 使用 fake prepay_id 占位，需对接微信支付统一下单 API 获取真实 prepay_id", "大", "商户号开通", "services/menu.py:251"),

    ("P1", "🛡️ 生产安全加固", "SECRET_KEY 生产化", "config.py 中未设置 SECRET_KEY 时自动生成随机密钥（重启失效），生产必须持久设置", "小", "部署前", "config.py:15-21"),
    ("P1", "🛡️ 生产安全加固", "限流升 Redis", "当前 per-user AI 限流用内存 dict (ai.py)，gunicorn 多 worker 下失效，需换 Redis 存储", "中", "Redis 服务", "services/ai.py"),
    ("P1", "🛡️ 生产安全加固", ".env.prod 完整化", "docker-compose.yml 依赖 .env 文件注入 SECRET_KEY/WECHAT_* 等，需准备生产 .env 模板", "小", "配置值收集完成", "docker-compose.yml:40-53"),
    ("P1", "🛡️ 生产安全加固", "PostgreSQL 迁移测试", "当前开发用 SQLite，生产用 PostgreSQL。需测试数据迁移 + Alembic schema", "中", "PostgreSQL 可用", "models.py:448"),

    ("P1", "📊 监控完善", "美团/抖音监控接入", "services/monitor.py:120-121 中 美团/抖音 标记为 skip，需补充搜索能力", "中", "opencli 适配器或 API", "services/monitor.py:120-121"),
    ("P1", "📊 监控完善", "OTA 订单同步", "携程/美团/飞猪的订单自动同步到本地 bookings 表（目前无此功能）", "大", "OTA 开放平台接入", "新功能"),

    ("P1", "🗺️ 旅游/美食数据", "山纪专属路线/美食", "seed_data.py 中 travel_routes 和 food_recommends 全部 bnb_id='guishu'，山纪需要茶园周边路线 + 山纪附近美食", "中", "山纪店主提供周边信息", "seed_data.py"),
    ("P1", "🗺️ 旅游/美食数据", "东林外专属路线/美食", "seed_data.py 中 travel_routes 和 food_recommends 全部 bnb_id='guishu'，东林外需要东林寺周边路线 + 附近美食", "中", "东林外店主提供周边信息", "seed_data.py"),
    ("P1", "🗺️ 旅游/美食数据", "山纪/东林外预览页生成", "preview/ 目录下 preview 页面只有归墅数据，需生成三店数据版本或动态化", "中", "种子数据完成后", "preview/*.html"),

    ("P2", "🎨 小程序完善", "小程序服务器域名白名单", "微信小程序后台配置 request 合法域名（需 HTTPS），否则正式版无法请求", "小", "域名部署 HTTPS", "小程序后台"),
    ("P2", "🎨 小程序完善", "小程序 Chat 页对接正式端点", "miniapp/pages/chat/chat.js 调用 /api/simulate-chat（开发用），生产需改为真实消息端点", "中", "微信消息接入完成", "miniapp/pages/chat/chat.js:34"),
    ("P2", "🎨 小程序完善", "小程序提交审核", "代码完成度约 70%，需补齐 shanji/donglinwai 数据后提交微信审核（1-7天）", "中", "数据补齐 + API 就绪", "miniapp/"),

    ("P2", "📈 运营增强", "Redis 缓存层", "热门页面（房型/菜单/攻略）加 Redis 缓存，减少数据库查询", "中", "Redis 服务", "新功能"),
    ("P2", "📈 运营增强", "AI 对话质量持续优化", "用户反馈收集 + prompt 迭代 + 多轮对话效果跟踪", "持续", "上线后数据积累", "services/ai.py"),
    ("P2", "📈 运营增强", "运营数据分析看板", "入住率/收入/客源渠道/口碑趋势 可视化看板（新页面/API）", "大", "数据积累后", "新功能"),
    ("P2", "📈 运营增强", "小红书/抖音自动发布", "通过 opencli 定时发布营销内容到小红书/抖音", "中", "opencli 可用 + 内容储备", "新功能"),
    ("P2", "📈 运营增强", "多语言支持", "面向外国游客的中英双语版本（庐山有国际游客）", "中", "模板改造", "templates/*.html"),
    ("P2", "📈 运营增强", "CI/CD 自动化部署", "GitHub Actions / Gitee → 自动测试 → Docker 构建 → 服务器部署", "中", "服务器就绪", "新功能"),
]

# ── Sheet 1: 全部待办按优先级 ──
ws1 = wb.active
ws1.title = "全部待办事项"

# Title row
ws1.merge_cells('A1:H1')
title_cell = ws1['A1']
title_cell.value = "云上·印象集 — 全部待办事项清单 (2026-06-23)"
title_cell.font = Font(name='Microsoft YaHei', size=14, bold=True, color='2a3d33')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 35

# Headers
headers = ['优先级', '分类', '任务', '详细说明', '工作量', '前置依赖', '涉及文件', '状态']
col_widths = [8, 22, 25, 55, 8, 22, 35, 8]
for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    cell = ws1.cell(row=2, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border
    ws1.column_dimensions[get_column_letter(i)].width = w

# Data rows
p_fills = {'P0': p0_fill, 'P1': p1_fill, 'P2': p2_fill, 'P3': p3_fill}
for idx, (priority, cat, task, detail, effort, dep, fpath) in enumerate(tasks, 3):
    row_data = [priority, cat, task, detail, effort, dep, fpath, '待处理']
    for col, val in enumerate(row_data, 1):
        cell = ws1.cell(row=idx, column=col, value=val)
        cell.font = bold_font if col <= 2 else normal_font
        cell.alignment = wrap if col == 4 else (center if col in [1, 5, 8] else Alignment(vertical='center'))
        cell.border = thin_border
        if priority in p_fills:
            cell.fill = p_fills[priority]
    ws1.row_dimensions[idx].height = 30

# Freeze panes
ws1.freeze_panes = 'A3'
ws1.auto_filter.ref = f'A2:H{len(tasks)+2}'

# ── Sheet 2: 山纪/东林外信息缺口专项 ──
ws2 = wb.create_sheet("山纪&东林外信息缺口")

ws2.merge_cells('A1:G1')
t2 = ws2['A1']
t2.value = "山纪 & 东林外 — 信息完善需求清单"
t2.font = Font(name='Microsoft YaHei', size=14, bold=True, color='2a3d33')
t2.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 35

headers2 = ['分类', '信息项', '归墅(现有)', '山纪(现状)', '东林外(现状)', '需收集什么', '优先级']
widths2 = [16, 14, 28, 28, 28, 32, 8]
for i, (h, w) in enumerate(zip(headers2, widths2), 1):
    cell = ws2.cell(row=2, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border
    ws2.column_dimensions[get_column_letter(i)].width = w

info_gaps = [
    ("基础信息", "详细地址", "庐山风景名胜区大林沟路27号", "XX路XX号 ❌", "XX路XX号 ❌", "真实门牌号/路口描述/地标", "P0"),
    ("基础信息", "联系电话", "16607927666", "待填写 ❌", "待填写 ❌", "前台/管家手机号", "P0"),
    ("基础信息", "GPS 坐标", "29.5568, 115.9797 (精确)", "29.5600, 115.9850 (估算) ⚠️", "29.5500, 115.9700 (估算) ⚠️", "百度/高德取点或实地标记", "P0"),
    ("基础信息", "民宿描述", "U型三层山居小院，11间精品客房...", "茶园环绕，品茗观山... (太短) ⚠️", "东林寺旁，禅意疗愈之所... (太短) ⚠️", "特色亮点/建筑风格/房间数/独特卖点", "P0"),
    ("基础信息", "主题色", "#4a7c59 (山居绿)", "#8B6914 (茶金)", "#7B8DAD (禅意蓝)", "已设定，可确认是否需要调整", "P2"),
    ("客房", "房型列表", "8 种房型 (单人间~套房)", "无 ❌", "无 ❌", "房型名/价格/原价/面积/床型/容量/景观/设施", "P0"),
    ("客房", "房间照片", "local_data/images/ 有 40+ 张", "无 ❌", "无 ❌", "每个房型至少 3-5 张高质量照片", "P0"),
    ("客房", "房间数量", "每种房型 total_count 已设定", "无 ❌", "无 ❌", "每种房型各有几间", "P0"),
    ("餐饮", "菜单", "21 道菜品 (咖啡/简餐/甜品)", "无 ❌", "无 ❌", "是否有餐饮服务？菜品/价格/分类", "P0"),
    ("餐饮", "特色饮食", "咖啡简餐为主", "茶园茶点？茶餐？", "素食？禅修餐？", "山纪茶相关饮食 / 东林外禅意饮食", "P1"),
    ("服务", "快捷服务", "14 项 (送水/加床/叫车等)", "无 ❌", "无 ❌", "支持哪些客房服务？", "P0"),
    ("服务", "特色体验", "无专属特色", "茶园体验 (采茶/制茶/茶道) ✅ 已有 seed", "疗愈课程 (冥想/瑜伽/音疗) ✅ 已有 seed", "已有 seed_tea 和 seed_healing，需确认内容准确性", "P1"),
    ("周边", "旅游路线", "5 条 (牯岭街/三叠泉/五老峰/好汉坡/东线)", "无 ❌", "无 ❌", "山纪周边茶园路线 / 东林外周边东林寺路线", "P1"),
    ("周边", "美食推荐", "8 家 (石牛酒家/847别墅/望庐说等)", "无 ❌", "无 ❌", "山纪附近餐馆 / 东林外附近餐馆", "P1"),
    ("周边", "娱乐休闲", "已整理 庐山牯岭镇娱乐休闲指南.md", "共用牯岭镇指南 ✅", "共用牯岭镇指南 ✅", "两家距牯岭镇距离不同，可能需要独立交通说明", "P1"),
    ("微信", "公众号 Token", "WECHAT_TOKEN_GS 环境变量", "WECHAT_TOKEN_SJ 环境变量", "WECHAT_TOKEN_DLW 环境变量", "每个公众号独立的 Token 值", "P1"),
    ("微信", "公众号 AppID/Secret", "WECHAT_APP_ID_GS/SECRET_GS", "WECHAT_APP_ID_SJ/SECRET_SJ", "WECHAT_APP_ID_DLW/SECRET_DLW", "每个公众号独立的 AppID/Secret", "P1"),
    ("预订", "OTA 平台链接", "已生成搜索链接", "已生成搜索链接 ✅", "已生成搜索链接 ✅", "三家店在各 OTA 的精确链接（如有独立页面）", "P2"),
]

for idx, row_data in enumerate(info_gaps, 3):
    for col, val in enumerate(row_data, 1):
        cell = ws2.cell(row=idx, column=col, value=val)
        cell.font = bold_font if col in [1, 2] else normal_font
        cell.alignment = wrap if col in [3, 4, 5, 6] else (center if col == 7 else Alignment(vertical='center'))
        cell.border = thin_border
        if "❌" in str(val):
            cell.fill = p0_fill
        elif "⚠️" in str(val):
            cell.fill = p1_fill
    ws2.row_dimensions[idx].height = 32

ws2.freeze_panes = 'A3'
ws2.auto_filter.ref = f'A2:G{len(info_gaps)+2}'

# ── Sheet 3: 按分类汇总 ──
ws3 = wb.create_sheet("按分类统计")

ws3.merge_cells('A1:D1')
t3 = ws3['A1']
t3.value = "按分类统计"
t3.font = Font(name='Microsoft YaHei', size=14, bold=True, color='2a3d33')
t3.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 30

h3 = ['分类', '待办数', '已完成/已有', '备注']
for i, h in enumerate(h3, 1):
    cell = ws3.cell(row=2, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

stats = [
    ("🏠 山纪/东林外·基础信息", 10, 0, "全部待收集"),
    ("🛏️ 山纪·客房/菜单", 3, 0, "茶树数据已有，客房/菜单空白"),
    ("🏕️ 东林外·客房/菜单", 3, 0, "疗愈数据已有，客房/菜单空白"),
    ("☁️ 域名与服务器", 9, 0, "Docker/Nginx 配置已就绪，缺服务器"),
    ("🔗 微信接入", 4, 0, "代码已实现，等待公众号认证+支付商户号"),
    ("🛡️ 生产安全加固", 4, 1, "基本框架已有，需 Redis + .env.prod"),
    ("📊 监控完善", 2, 1, "4 平台运转正常，美团/抖音/OTA 未接入"),
    ("🗺️ 旅游/美食数据", 3, 1, "归墅数据完整，山纪/东林外空白"),
    ("🎨 小程序完善", 3, 0, "前端 UI 完整，数据+API 待补齐"),
    ("📈 运营增强", 6, 1, "AI 对话/周报已运转，其余需数据积累"),
]

for idx, (cat, count, done, note) in enumerate(stats, 3):
    for col, val in enumerate([cat, count, done, note], 1):
        cell = ws3.cell(row=idx, column=col, value=val)
        cell.font = bold_font if col == 1 else normal_font
        cell.alignment = center if col in [2, 3] else Alignment(vertical='center')
        cell.border = thin_border
    ws3.row_dimensions[idx].height = 22

# Totals
total_row = len(stats) + 3
total_tasks = sum(s[1] for s in stats)
total_done = sum(s[2] for s in stats)
ws3.cell(row=total_row, column=1, value="合计").font = bold_font
ws3.cell(row=total_row, column=1).border = thin_border
ws3.cell(row=total_row, column=2, value=total_tasks).font = bold_font
ws3.cell(row=total_row, column=2).alignment = center
ws3.cell(row=total_row, column=2).border = thin_border
ws3.cell(row=total_row, column=3, value=total_done).font = bold_font
ws3.cell(row=total_row, column=3).alignment = center
ws3.cell(row=total_row, column=3).border = thin_border
ws3.cell(row=total_row, column=4, value=f"总计 {total_tasks} 项待办，{total_done} 项已完成").font = bold_font
ws3.cell(row=total_row, column=4).border = thin_border

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 12
ws3.column_dimensions['C'].width = 18
ws3.column_dimensions['D'].width = 45
ws3.freeze_panes = 'A3'

# ── Sheet 4: 数据收集模板（给店主填的）──
ws4 = wb.create_sheet("店主信息收集模板")

ws4.merge_cells('A1:F1')
t4 = ws4['A1']
t4.value = "山纪 & 东林外 — 信息收集表（请店主填写）"
t4.font = Font(name='Microsoft YaHei', size=14, bold=True, color='2a3d33')
t4.alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 35

# Section A: Basic Info
row = 3
ws4.merge_cells(f'A{row}:F{row}')
ws4.cell(row=row, column=1, value="A. 基础信息").font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
ws4.cell(row=row, column=1).fill = PatternFill(start_color='4a7c59', end_color='4a7c59', fill_type='solid')
for c in range(1, 7):
    ws4.cell(row=row, column=c).border = thin_border

row = 4
for i, h in enumerate(['信息项', '说明', '山纪 请填写 →', '', '东林外 请填写 →', ''], 1):
    ws4.cell(row=row, column=i, value=h).font = bold_font
    ws4.cell(row=row, column=i).border = thin_border
    ws4.cell(row=row, column=i).alignment = center

basic_fields = [
    ("民宿全称", "营业执照上的正式名称", "", ""),
    ("详细地址", "完整的门牌号/路口描述（例：庐山风景名胜区大林沟路27号）", "", ""),
    ("联系电话", "前台或管家手机号（客人可见）", "", ""),
    ("GPS纬度", "百度/高德地图取点（例：29.5568）", "", ""),
    ("GPS经度", "百度/高德地图取点（例：115.9797）", "", ""),
    ("民宿简介", "100-200字的特色描述：建筑风格、独特卖点、氛围、房间数", "", ""),
    ("主题色", "建议的网页主题色（例：#4a7c59 山居绿）", "", ""),
]

for i, (field, desc, sj, dlw) in enumerate(basic_fields):
    r = row + 1 + i
    ws4.cell(row=r, column=1, value=field).font = bold_font
    ws4.cell(row=r, column=2, value=desc).font = normal_font
    ws4.cell(row=r, column=3, value=sj).font = normal_font
    ws4.cell(row=r, column=5, value=dlw).font = normal_font
    for c in range(1, 7):
        ws4.cell(row=r, column=c).border = thin_border
        ws4.cell(row=r, column=c).alignment = wrap
    ws4.row_dimensions[r].height = 28
    # merge col 3-4 for 山纪, 5-6 for 东林外
    ws4.merge_cells(f'C{r}:D{r}')
    ws4.merge_cells(f'E{r}:F{r}')

# Section B: Rooms
row = row + len(basic_fields) + 2
ws4.merge_cells(f'A{row}:F{row}')
ws4.cell(row=row, column=1, value="B. 客房信息（每种房型一行，自行增加行数）").font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
ws4.cell(row=row, column=1).fill = PatternFill(start_color='4a7c59', end_color='4a7c59', fill_type='solid')
for c in range(1, 7):
    ws4.cell(row=row, column=c).border = thin_border

row += 1
room_headers = ['房型名称', '价格(元/晚)', '原价', '面积(m²)', '床型', '容量(人)', '景观', '设施(逗号分隔)', '房量(间)', '所属民宿']
for i, h in enumerate(room_headers, 1):
    ws4.cell(row=row, column=i, value=h).font = bold_font
    ws4.cell(row=row, column=i).border = thin_border
    ws4.cell(row=row, column=i).alignment = wrap
    ws4.cell(row=row, column=i).fill = PatternFill(start_color='d4e0db', end_color='d4e0db', fill_type='solid')

# Example rows for rooms
room_examples = [
    ["示例: 茶园大床房", "588", "688", "25", "1.8m大床", "2", "茶园山景", "全景窗,地暖,茶具套装,高速WiFi,迷你吧", "2", "shanji"],
    ["示例: 禅修单人房", "388", "488", "15", "1.5m双人床", "1", "庭院", "观景窗,地暖,书桌,高速WiFi", "2", "donglinwai"],
]
for i, ex in enumerate(room_examples):
    r = row + 1 + i
    for c, v in enumerate(ex, 1):
        ws4.cell(row=r, column=c, value=v).font = normal_font
        ws4.cell(row=r, column=c).border = thin_border
        ws4.cell(row=r, column=c).alignment = wrap
    ws4.row_dimensions[r].height = 25

# Section C: Menu
row = row + 4
ws4.merge_cells(f'A{row}:F{row}')
ws4.cell(row=row, column=1, value="C. 菜单信息（如有餐饮服务）").font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
ws4.cell(row=row, column=1).fill = PatternFill(start_color='4a7c59', end_color='4a7c59', fill_type='solid')
for c in range(1, 7):
    ws4.cell(row=row, column=c).border = thin_border

row += 1
menu_headers = ['菜品名称', '分类', '价格(元)', '描述', '所属民宿']
for i, h in enumerate(menu_headers, 1):
    ws4.cell(row=row, column=i, value=h).font = bold_font
    ws4.cell(row=row, column=i).border = thin_border
    ws4.cell(row=row, column=i).fill = PatternFill(start_color='d4e0db', end_color='d4e0db', fill_type='solid')

menu_examples = [
    ["示例: 庐山云雾茶", "茶饮", "68", "中国十大名茶，山泉水冲泡，配茶点", "shanji"],
    ["示例: 素斋套餐", "简餐", "48", "时令蔬菜 + 米饭 + 汤", "donglinwai"],
]
for i, ex in enumerate(menu_examples):
    r = row + 1 + i
    for c, v in enumerate(ex, 1):
        ws4.cell(row=r, column=c, value=v).font = normal_font
        ws4.cell(row=r, column=c).border = thin_border
        ws4.cell(row=r, column=c).alignment = wrap
    ws4.row_dimensions[r].height = 25

# Section D: Services
row = row + 4
ws4.merge_cells(f'A{row}:F{row}')
ws4.cell(row=row, column=1, value="D. 快捷服务（客人可通过微信/小程序请求的服务）").font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
ws4.cell(row=row, column=1).fill = PatternFill(start_color='4a7c59', end_color='4a7c59', fill_type='solid')
for c in range(1, 7):
    ws4.cell(row=row, column=c).border = thin_border

row += 1
svc_headers = ['服务名称', '分类', '所属民宿', '备注']
for i, h in enumerate(svc_headers, 1):
    ws4.cell(row=row, column=i, value=h).font = bold_font
    ws4.cell(row=row, column=i).border = thin_border
    ws4.cell(row=row, column=i).fill = PatternFill(start_color='d4e0db', end_color='d4e0db', fill_type='solid')

svc_examples = [
    ["送矿泉水", "housekeeping", "shanji", ""],
    ["加床", "housekeeping", "donglinwai", ""],
    ["茶道体验预约", "frontdesk", "shanji", "山纪特色"],
    ["冥想课程预约", "frontdesk", "donglinwai", "东林外特色"],
]
for i, ex in enumerate(svc_examples):
    r = row + 1 + i
    for c, v in enumerate(ex, 1):
        ws4.cell(row=r, column=c, value=v).font = normal_font
        ws4.cell(row=r, column=c).border = thin_border
        ws4.cell(row=r, column=c).alignment = wrap
    ws4.row_dimensions[r].height = 25

# Section E: Surrounding
row = row + 5
ws4.merge_cells(f'A{row}:F{row}')
ws4.cell(row=row, column=1, value="E. 周边推荐（路线 & 美食）").font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
ws4.cell(row=row, column=1).fill = PatternFill(start_color='4a7c59', end_color='4a7c59', fill_type='solid')
for c in range(1, 7):
    ws4.cell(row=row, column=c).border = thin_border

row += 1
ws4.merge_cells(f'A{row}:F{row}')
ws4.cell(row=row, column=1, value="周边旅游路线（每条一行）：路线名称 / 距离 / 交通方式 / 大约用时 / 适合季节 / 难度 / 简要描述 / GPS坐标 / 所属民宿").font = normal_font
for c in range(1, 7):
    ws4.cell(row=row, column=c).border = thin_border

row += 1
ws4.merge_cells(f'A{row}:F{row}')
ws4.cell(row=row, column=1, value="周边美食推荐（每家一行）：餐厅名称 / 类型 / 人均 / 推荐菜 / 距离 / GPS坐标 / 所属民宿").font = normal_font
for c in range(1, 7):
    ws4.cell(row=row, column=c).border = thin_border

# Column widths for sheet 4
ws4.column_dimensions['A'].width = 16
ws4.column_dimensions['B'].width = 50
ws4.column_dimensions['C'].width = 18
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 18
ws4.column_dimensions['F'].width = 18

# Save
output_path = r'D:\first-cc\reports\待办事项清单_20260623.xlsx'
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"✅ 已生成: {output_path}")
print(f"   Sheet 1: 全部待办事项 ({len(tasks)} 项)")
print(f"   Sheet 2: 山纪&东林外信息缺口 ({len(info_gaps)} 项)")
print(f"   Sheet 3: 按分类统计 ({len(stats)} 类)")
print(f"   Sheet 4: 店主信息收集模板")
